"""This script is used to compare the LLM output by comparing replies in English and not for the Italian, Spanish, and German languages."""

import os
import sys

import pandas as pd

# =========================================================
# Input correct results from command line
try:
    input_dir = str(sys.argv[1])
except IndexError:
    print("No input directory provided. Using default.")
    input_dir = "/Users/leonardo/git/malco/in_manual_curation"
file = os.path.join(input_dir, "prompts", "correct_results.tsv")
try:
    output_dir = str(sys.argv[2])
except IndexError:
    print("No output directory provided. Using default.")
    output_dir = "/Users/leonardo/git/malco/out_manual_curation"
try:
    result_dir = str(sys.argv[3])
except IndexError:
    print("No result directory provided. Using default.")
    result_dir = "/Users/leonardo/git/malco/analysis_out"

result_suffix_to_remove = "_en-prompt.txt.result"
remove_result = len(result_suffix_to_remove)
prompt_filename_suffix_to_remove = "_en-prompt.txt"
remove_std_suffix = len(prompt_filename_suffix_to_remove)

# Mapping each label to its correct term
answers = pd.read_csv(file, sep="\t", header=None, names=["description", "term", "label"])
cres = answers.set_index("label").to_dict(orient="index")
answers_dict = {k[:-remove_std_suffix]: v for k, v in cres.items()}

langs = ["it", "es", "de"]
# Define a dataframe with 100 lines and 9 columns
df = pd.DataFrame(
    columns=[
        "PMID",
        "correct_label",
        "correct_OMIM_id",
        "it_dx",
        "es_dx",
        "de_dx",
        "it_rank",
        "es_rank",
        "de_rank",
    ]
)

# Load ungrounded results
ungrounded_dir = os.path.join(output_dir, "raw_results", "multilingual")
grounded_dir = os.path.join(output_dir, "multilingual")


# Loop over directories in ungrounded_dir and use those cotaining "no_en" in their name

for lang in langs:
    dir_path = os.path.join(ungrounded_dir, lang + "_no_en", "differentials_by_file")
    if os.path.exists(dir_path):
        print("Found directory:", dir_path)
        # Loop over files in the directory
        for file in os.listdir(dir_path):
            file_path = os.path.join(dir_path, file)
            # Read the content of the file and populate the dictionary
            with open(file_path, "r") as f:
                file_content = f.read()
            file_key = file[:-remove_result]  # Remove the last n characters from the filename
            if file_key not in df["PMID"].values:
                # Add a new row with the file_key in the PMID column and file content in the lang+"_no_en" column
                df = pd.concat(
                    [df, pd.DataFrame({"PMID": [file_key], lang + "_dx": [file_content]})],
                    ignore_index=True,
                )
            else:
                # Update the existing row in the lang+"_no_en" column
                df.loc[df["PMID"] == file_key, lang + "_dx"] = file_content

    dir_path = os.path.join(grounded_dir, lang + "_w_en")
    if os.path.exists(dir_path):
        print("Found directory:", dir_path)
        fulldf = pd.read_csv(os.path.join(dir_path, "full_df_results.tsv"), sep="\t")
        # Group by label and iterate over the groups
        for label, group in fulldf.groupby("label"):
            label = label[:-remove_std_suffix]  # Remove the last n characters from the label
            # Check if the label exists in df
            if label not in df["PMID"].values:
                print(f"Label {label} not found in df.")
                continue
            # If any item in column "is_correct" is True, set the corresponding "rank" value in df[lang+"_rank"]
            if group["is_correct"].any():
                # Get the rank of the first correct item
                rank = group.loc[group["is_correct"], "rank"].values[0]
                # Update the corresponding row in df
                df.loc[df["PMID"] == label, lang + "_rank"] = rank
            else:
                # If no correct item, set rank to NaN
                df.loc[df["PMID"] == label, lang + "_rank"] = float("NaN")

# Add the correct MONDO ID and description to the dataframe
df["correct_OMIM_id"] = df["PMID"].map(
    lambda x: answers_dict[x]["term"] if x in answers_dict else None
)
df["correct_label"] = df["PMID"].map(
    lambda x: answers_dict[x]["description"] if x in answers_dict else None
)

# ==========================================================================================================================
# ==========================================================================================================================
# Some excel magic to make the output more readable
# =============================================================
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Save the DataFrame to an Excel file
excel_path = os.path.join(result_dir, "replies2curate.xlsx")
df.to_excel(excel_path, index=False)

# Open the Excel file and enable "Wrap Text" for all cells
wb = load_workbook(excel_path)
ws = wb.active

for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str) and "\n" in cell.value:
            cell.alignment = Alignment(wrap_text=True)

# Automatically adjust column widths based on content
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)  # Get the column letter (e.g., A, B, C)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    # Set the column width
    ws.column_dimensions[col_letter].width = max_length
    # For columns D, E, F take a smaller width
    if col_letter in ["D", "E", "F"]:
        ws.column_dimensions[col_letter].width = max_length / 2

# Automatically adjust row height based on content
for row in ws.iter_rows():
    max_height = 0
    for cell in row:
        if cell.value and isinstance(cell.value, str) and "\n" in cell.value:
            max_height = max(max_height, cell.value.count("\n") + 1)
    # Set the row height (add a small buffer for better readability)
    ws.row_dimensions[row[0].row].height = max_height * 15  # Adjust the multiplier as needed

# Save the updated Excel file
wb.save(excel_path)
# ==========================================================================================================================
# ==========================================================================================================================
# Also save the dataframe to a TSV file
# =============================================================
# Escape newlines explicitly for all string fields in the DataFrame
df = df.applymap(lambda x: x.replace("\n", "\\n") if isinstance(x, str) else x)
df.to_csv(os.path.join(result_dir, "replies2curate.tsv"), sep="\t", index=False)
